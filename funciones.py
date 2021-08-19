from openpyxl import load_workbook

filename = './data.xlsx'

# 1: Read
def read_records(status: str = 'Todas')->list:
  records = []
  file = load_workbook(filename)
  sheet = file['Hoja1']
  rows = sheet.rows
  for i,row in enumerate(rows):
    if (i == 0):
      continue
    record = convert_row_to_dict(row)
    records.append(record)

  if (status != 'Todas'):
    return [record for record in records if (record['Estado'] == status)]
  file.close()
  return records

def convert_row_to_dict(row: tuple)-> dict:
  dictionary = {
    'Id': row[0].value,
    'Tarea': row[1].value,
    'Descripcion': row[2].value,
    'Estado': row[3].value,
    'Fecha inicio': row[4].value,
    'Fecha finalizacion': row[5].value,
  }
  return dictionary

# TODO: Implementar filtrado
# records = read_records()
# for record in records:
#   print('--------------------------')
#   print(f"id: {record['Id']}\nTarea: {record['Tarea']}\nDescripcion: {record['Descripcion']}")

def update_row_with_dict(row: tuple, record: dict):
  if (record.get('Tarea')): row[1].value = record.get('Tarea')
  if (record.get('Descripcion')): row[2].value = record.get('Descripcion')
  if (record.get('Estado')): row[3].value = record.get('Estado')
  if (record.get('Fecha inicio')): row[4].value = record.get('Fecha inicio')
  if (record.get('Fecha finalizacion')): row[5].value = record.get('Fecha finalizacion')

# 2: Create
def add_record(record: dict):

  recordId = record['Id']
  file = load_workbook(filename)
  sheet = file['Hoja1']
  row_idx = sheet.max_row + 1
  cell = sheet.cell(row=row_idx, column=1)
  cell.value = recordId

  for row in sheet.rows:
    if (row[0].value == recordId):
      update_row_with_dict(row, record)

  file.save(filename)
  file.close()

# add_record({
#   'Id': 3,
#   'Tarea': 'Caminar',
#   'Descripcion': 'Ir a caminar'
# })

# 3: Delete
def delete_record(id: int):
  file = load_workbook(filename)
  sheet = file['Hoja1']
  row_to_delete_idx = None
  for row in sheet.rows:
    first_cell = row[0]
    if (first_cell.value == id):
      row_to_delete_idx = first_cell.row
      sheet.delete_rows(idx=row_to_delete_idx, amount=1)
  
  if (row_to_delete_idx != None):
    try:
      file.save(filename)
    except:
      print('El archivo no puede ser modificado en este momento')
  else:
    print(f'No se encontró tarea con id = {id}')
  file.close()

# delete_record(3)

# 4: Update
def update_record(id: int, record: dict):

  file = load_workbook(filename)
  sheet = file['Hoja1']
  found = False
  for row in sheet.rows:
    first_cell = row[0]
    if (first_cell.value == id):
      found = True
      update_row_with_dict(row, record)
  if (found):
    file.save(filename)
  else:
    print(f'No se encontró tarea con id = {id}')
  file.close()

# update_record(3, {
#   # 'Id': 3,
#   'Tarea': 'Salir de paseo',
#   'Descripcion': 'Ir a conocer algun lugar nuevo'
# })

def print_menu_and_ask_for_selection()-> str:
  print('--------Menu---------')
  print('1. Consultar')
  print('2. Crear una tarea')
  print('3. Actualizar una tarea')
  print('4. Eliminar una tarea')
  print('---------------------')
  return input('Ingrese su selección: ')

def ask_for_status()->str:
  options = {
    1: 'Todas',
    2: 'Pendiente',
    3: 'En ejecución',
    4: 'Finalizada',
  }
  print('Estados:')
  for key, value in options.items():
    print(f'{key}: {value}')
  seleccion = int(input('Ingrese el estado: '))
  return options[seleccion]

def print_records(records: list[dict]):
  for record in records:
    print('------------------')
    print(f"Id: {record['Id']}")
    print(f"Tarea: {record['Tarea']}")
    print(f"Descripcion: {record['Descripcion']}")
    print(f"Estado: {record['Estado']}")
    print(f"Fecha inicio: {record['Fecha inicio']}")
    print(f"Fecha finalizacion: {record['Fecha finalizacion']}")